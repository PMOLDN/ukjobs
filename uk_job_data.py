"""Utilities for building UK occupation data from ONS workbooks."""

from __future__ import annotations

import csv
import json
import re
from pathlib import Path

from openpyxl import load_workbook

DEFAULT_WORKBOOK = "employmentby4digitsoc1digitindustryatosuk20212024final.xlsx"
DEFAULT_DESCRIPTION_WORKBOOK = "soc2020_unit_groups.xlsx"
DEFAULT_PAY_WORKBOOK = (
    "Data/Raw/pay/ashetable142024provisional/"
    "PROV - Occupation SOC20 (4) Table 14.7a   Annual pay - Gross 2024.xlsx"
)
DEFAULT_REGIONAL_CSV = "Data/Raw/regional/731911681617723.csv"
SOC2020_DESCRIPTION_DOWNLOAD_URL = (
    "https://www.ons.gov.uk/file?uri=%2Fmethodology%2Fclassificationsandstandards"
    "%2Fstandardoccupationalclassificationsoc%2Fsoc2020%2Fsoc2020volume1structure"
    "anddescriptionsofunitgroups%2Fsoc2020volume1structureanddescriptionofunitgroups"
    "excel03122025.xlsx"
)
YEARS = ("2021", "2022", "2023", "2024")
DATASET_ID = "uk-ons-aps"
SOURCE_NAME = "Office for National Statistics Annual Population Survey"
PAY_SOURCE_NAME = "Office for National Statistics Annual Survey of Hours and Earnings Table 14"
REGIONAL_SOURCE_NAME = "Nomis Annual Population Survey regional occupation (SOC2020)"
DESCRIPTION_SOURCE_NAME = (
    "Office for National Statistics SOC 2020 Volume 1: Structure and descriptions of unit groups"
)
REGIONAL_PERIOD = "Jan 2024-Dec 2024"
REGIONAL_METADATA_KEYS = {
    "Area Type  :",
    "Area Name  :",
    "Permanent/non-permanent:",
    "Full/Part-time:",
    "Employee/Self-employed:",
    "Sex        :",
}
DEFAULT_REGION = "London"

MAJOR_GROUP_LABELS = {
    "1": "Managers, directors and senior officials",
    "2": "Professional occupations",
    "3": "Associate professional and technical occupations",
    "4": "Administrative and secretarial occupations",
    "5": "Skilled trades occupations",
    "6": "Caring, leisure and other service occupations",
    "7": "Sales and customer service occupations",
    "8": "Process, plant and machine operatives",
    "9": "Elementary occupations",
}


def clean_text(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip()


def slugify(text: str) -> str:
    text = text.lower().replace("&", " and ")
    text = re.sub(r"[^a-z0-9]+", "-", text)
    return re.sub(r"-{2,}", "-", text).strip("-")


def parse_industry_header(value: object) -> dict[str, str] | None:
    if value is None:
        return None
    text = clean_text(str(value))
    match = re.match(r"^\d+\s+([A-S])\s+(.+)$", text)
    if not match:
        return None
    return {"code": match.group(1), "label": match.group(2)}


def parse_occupation_label(value: object) -> tuple[str, str] | None:
    if value is None:
        return None
    text = clean_text(str(value))
    if text.lower() == "total":
        return None
    match = re.match(r"^(\d{4})\s+(.+)$", text)
    if not match:
        return None
    return match.group(1), match.group(2)


def parse_jobs_cell(value: object) -> int | None:
    if isinstance(value, (int, float)):
        return int(round(value))
    return None


def parse_numeric_cell(value: object) -> int | None:
    if isinstance(value, (int, float)):
        return int(round(value))
    text = clean_text(str(value or "")).replace(",", "")
    if not text or text in {"*", "~", "!", "x", "X", "..", ":", "-"}:
        return None
    try:
        return int(round(float(text)))
    except ValueError:
        return None


def is_suppressed(value: object) -> bool:
    return isinstance(value, str) and value.strip() == "*"


def is_empty(value: object) -> bool:
    return value is None or (isinstance(value, str) and value.strip() in {"", "-"})


def describe_trend(trend_pct: float | None) -> str:
    if trend_pct is None:
        return "No comparable trend"
    if trend_pct <= -15:
        return "Sharp decline"
    if trend_pct <= -5:
        return "Declining"
    if trend_pct < 5:
        return "Broadly flat"
    if trend_pct < 15:
        return "Growing"
    return "Rapid growth"


def normalize_blank(value: object) -> str:
    text = clean_text(str(value or ""))
    return "" if text == "<blank>" else text


def normalize_code(value: object) -> str:
    if isinstance(value, (int, float)):
        return str(int(value))
    text = clean_text(str(value or ""))
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return text


def split_tilde_bullets(value: object) -> list[str]:
    text = normalize_blank(value)
    if not text:
        return []
    parts = []
    for chunk in re.split(r"(?:\n\s*\n|~)", text):
        cleaned = clean_text(chunk)
        if not cleaned:
            continue
        parts.append(cleaned)
    return parts


def format_number(value: int | None) -> str:
    if value is None:
        return "n/a"
    return f"{value:,}"


def format_percent(value: float | None) -> str:
    if value is None:
        return "n/a"
    sign = "+" if value > 0 else ""
    return f"{sign}{value:.1f}%"


def _industry_columns(sheet) -> list[tuple[int, dict[str, str]]]:
    columns: list[tuple[int, dict[str, str]]] = []
    started = False
    blank_run = 0

    for col in range(3, sheet.max_column + 1):
        header = parse_industry_header(sheet.cell(8, col).value)
        if header is None:
            if started:
                blank_run += 1
                if blank_run >= 3:
                    break
            continue
        started = True
        blank_run = 0
        columns.append((col, header))

    if not columns:
        raise ValueError(f"Could not find industry headers in sheet {sheet.title}")

    return columns


def _load_year(sheet, columns: list[tuple[int, dict[str, str]]]) -> dict[str, dict]:
    records: dict[str, dict] = {}
    max_col = max(col for col, _ in columns)
    col_to_industry = {col: industry for col, industry in columns}

    for values in sheet.iter_rows(min_row=9, max_row=sheet.max_row, max_col=max_col, values_only=True):
        parsed = parse_occupation_label(values[1] if len(values) > 1 else None)
        if parsed is None:
            if clean_text(str((values[1] if len(values) > 1 else "") or "")).lower() == "total":
                break
            continue

        soc_code, title = parsed
        total_jobs = 0
        suppressed_cells = 0
        missing_cells = 0
        breakdown = []

        for col, raw_value in enumerate(values, start=1):
            industry = col_to_industry.get(col)
            if industry is None:
                continue
            jobs = parse_jobs_cell(raw_value)
            if jobs is not None:
                total_jobs += jobs
                breakdown.append(
                    {
                        "code": industry["code"],
                        "label": industry["label"],
                        "jobs": jobs,
                    }
                )
            elif is_suppressed(raw_value):
                suppressed_cells += 1
            elif is_empty(raw_value):
                missing_cells += 1

        breakdown.sort(key=lambda item: item["jobs"], reverse=True)
        dominant = breakdown[0] if breakdown else None

        records[soc_code] = {
            "soc_code": soc_code,
            "title": title,
            "jobs": total_jobs,
            "suppressed_cells": suppressed_cells,
            "missing_cells": missing_cells,
            "industry_breakdown": breakdown,
            "top_industries": breakdown[:3],
            "dominant_industry_code": dominant["code"] if dominant else "",
            "dominant_industry": dominant["label"] if dominant else "",
            "dominant_industry_jobs": dominant["jobs"] if dominant else 0,
            "dominant_industry_share": (
                dominant["jobs"] / total_jobs if dominant and total_jobs else None
            ),
        }

    return records


def load_uk_pay_map(workbook_path: str = DEFAULT_PAY_WORKBOOK) -> dict[str, int | None]:
    path = Path(workbook_path)
    if not path.exists():
        return {}

    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook["All"]
    pay_map: dict[str, int | None] = {}

    for row in sheet.iter_rows(min_row=6, values_only=True):
        code = normalize_code(row[1] if len(row) > 1 else None)
        if not re.fullmatch(r"\d{4}", code):
            continue
        pay_map[code] = parse_numeric_cell(row[3] if len(row) > 3 else None)

    return pay_map


def parse_regional_parent_row(value: object) -> tuple[str, str] | None:
    text = clean_text(str(value or ""))
    match = re.match(r"^(\d{1,3})\s*:\s*(.+)$", text)
    if not match:
        return None
    return match.group(1), match.group(2)


def _sort_area_names(names: list[str]) -> list[str]:
    return sorted(
        names,
        key=lambda name: (0 if name == DEFAULT_REGION else 1, name.lower()),
    )


def load_uk_regional_metrics(
    csv_path: str = DEFAULT_REGIONAL_CSV,
    period: str = REGIONAL_PERIOD,
) -> dict[str, object]:
    path = Path(csv_path)
    if not path.exists():
        return {
            "available_areas": [],
            "default_area": "",
            "soc3_labels": {},
            "soc3_metrics": {},
        }

    blocks: dict[str, dict] = {}
    soc3_labels: dict[str, str] = {}
    current_meta: dict[str, str] = {}
    current_area_name = ""
    current_collecting = False
    current_value_index: int | None = None

    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        for row in reader:
            if not row:
                continue

            first = row[0].strip()
            if first == "annual population survey - regional - occupation (SOC2020) by sex by employment type":
                current_collecting = False
                current_value_index = None
                continue

            if len(row) >= 2 and row[0] in REGIONAL_METADATA_KEYS:
                current_meta[row[0]] = row[1].strip()
                continue

            if first == "Occupation (SOC2020)":
                current_value_index = row.index(period) if period in row else None
                current_collecting = (
                    current_value_index is not None
                    and current_meta.get("Permanent/non-permanent:") == "Total"
                    and current_meta.get("Full/Part-time:") == "Total"
                    and current_meta.get("Employee/Self-employed:") == "Total"
                    and current_meta.get("Sex        :") == "All persons"
                    and current_meta.get("Area Type  :") in {"countries", "regions"}
                )
                current_area_name = current_meta.get("Area Name  :", "")
                if current_collecting and current_area_name:
                    blocks.setdefault(
                        current_area_name,
                        {
                            "area_type": current_meta.get("Area Type  :", ""),
                            "total_jobs": None,
                            "codes": {},
                        },
                    )
                continue

            if not current_collecting or current_value_index is None or not current_area_name:
                continue

            if first == "Total":
                value = parse_numeric_cell(
                    row[current_value_index] if current_value_index < len(row) else None
                )
                blocks[current_area_name]["total_jobs"] = value
                continue

            parsed = parse_regional_parent_row(first)
            if parsed is None:
                continue

            code, label = parsed
            if len(code) != 3:
                continue

            value = parse_numeric_cell(
                row[current_value_index] if current_value_index < len(row) else None
            )
            blocks[current_area_name]["codes"][code] = value
            soc3_labels.setdefault(code, label)

    uk_block = blocks.get("United Kingdom")
    uk_total = uk_block.get("total_jobs") if uk_block else None
    available_areas = _sort_area_names(
        [
            name
            for name, block in blocks.items()
            if name != "United Kingdom"
            and block.get("area_type") in {"countries", "regions"}
            and block.get("total_jobs") is not None
        ]
    )

    soc3_metrics: dict[str, dict] = {}
    for code, label in soc3_labels.items():
        uk_jobs = uk_block["codes"].get(code) if uk_block else None
        jobs_by_area: dict[str, int | None] = {}
        lq_by_area: dict[str, float | None] = {}

        for area in available_areas:
            area_block = blocks[area]
            area_jobs = area_block["codes"].get(code)
            area_total = area_block.get("total_jobs")
            jobs_by_area[area] = area_jobs

            lq = None
            if (
                area_jobs is not None
                and area_total
                and uk_jobs is not None
                and uk_total
                and uk_jobs > 0
            ):
                area_share = area_jobs / area_total
                uk_share = uk_jobs / uk_total
                if uk_share > 0:
                    lq = round(area_share / uk_share, 3)

            lq_by_area[area] = lq

        soc3_metrics[code] = {
            "label": label,
            "jobs_by_area": jobs_by_area,
            "lq_by_area": lq_by_area,
        }

    return {
        "available_areas": available_areas,
        "default_area": DEFAULT_REGION if DEFAULT_REGION in available_areas else (available_areas[0] if available_areas else ""),
        "soc3_labels": soc3_labels,
        "soc3_metrics": soc3_metrics,
    }


def load_uk_records(workbook_path: str = DEFAULT_WORKBOOK) -> list[dict]:
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    columns = _industry_columns(workbook[YEARS[-1]])
    year_maps = {year: _load_year(workbook[year], columns) for year in YEARS}
    pay_map = load_uk_pay_map()
    regional_metrics = load_uk_regional_metrics()

    merged: list[dict] = []
    for soc_code in sorted({code for data in year_maps.values() for code in data}):
        latest = year_maps[YEARS[-1]].get(soc_code)
        fallback = latest or next(data[soc_code] for data in year_maps.values() if soc_code in data)
        jobs_by_year = {year: year_maps[year].get(soc_code, {}).get("jobs", 0) for year in YEARS}
        baseline = jobs_by_year["2021"]
        current = jobs_by_year["2024"]
        trend_pct = ((current - baseline) / baseline * 100) if baseline else None
        trend_change = (current - baseline) if baseline else None
        major_group_code = soc_code[0]
        minor_group_code = soc_code[:3]
        category_label = MAJOR_GROUP_LABELS.get(major_group_code, "Other occupations")
        regional_parent = regional_metrics["soc3_metrics"].get(minor_group_code, {})

        merged.append(
            {
                "dataset": DATASET_ID,
                "source": SOURCE_NAME,
                "source_file": Path(workbook_path).name,
                "title": fallback["title"],
                "slug": "",
                "url": "",
                "category": slugify(category_label),
                "category_label": category_label,
                "major_group_code": major_group_code,
                "major_group_label": category_label,
                "minor_group_code": minor_group_code,
                "minor_group_label": regional_metrics["soc3_labels"].get(minor_group_code, ""),
                "soc_code": soc_code,
                "num_jobs_2021": jobs_by_year["2021"],
                "num_jobs_2022": jobs_by_year["2022"],
                "num_jobs_2023": jobs_by_year["2023"],
                "num_jobs_2024": current,
                "trend_pct_2021_2024": round(trend_pct, 1) if trend_pct is not None else None,
                "trend_desc": describe_trend(trend_pct),
                "employment_change_2021_2024": trend_change,
                "dominant_industry_code": fallback["dominant_industry_code"],
                "dominant_industry": fallback["dominant_industry"],
                "dominant_industry_jobs": fallback["dominant_industry_jobs"],
                "dominant_industry_share": round(fallback["dominant_industry_share"], 4)
                if fallback["dominant_industry_share"] is not None
                else None,
                "suppressed_cells_2021": year_maps["2021"].get(soc_code, {}).get("suppressed_cells", 0),
                "suppressed_cells_2022": year_maps["2022"].get(soc_code, {}).get("suppressed_cells", 0),
                "suppressed_cells_2023": year_maps["2023"].get(soc_code, {}).get("suppressed_cells", 0),
                "suppressed_cells_2024": fallback["suppressed_cells"],
                "missing_cells_2024": fallback["missing_cells"],
                "industry_breakdown_2024": fallback["industry_breakdown"],
                "top_industries_2024": fallback["top_industries"],
                "median_pay_annual": pay_map.get(soc_code),
                "regional_parent_jobs_2024": regional_parent.get("jobs_by_area", {}),
                "regional_parent_lq_2024": regional_parent.get("lq_by_area", {}),
            }
        )

    _assign_unique_slugs(merged)
    return merged


def _assign_unique_slugs(records: list[dict]) -> None:
    seen: dict[str, int] = {}
    for record in records:
        base_slug = slugify(record["title"])
        if not base_slug:
            base_slug = f"occupation-{record['soc_code']}"
        count = seen.get(base_slug, 0) + 1
        seen[base_slug] = count
        record["slug"] = base_slug if count == 1 else f"{base_slug}-{record['soc_code']}"


def write_occupations_json(records: list[dict], path: str = "occupations.json") -> None:
    with open(path, "w", encoding="utf-8") as handle:
        json.dump(records, handle, indent=2)


def load_uk_descriptions(
    workbook_path: str = DEFAULT_DESCRIPTION_WORKBOOK,
) -> dict[str, dict]:
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    sheet = workbook["SOC2020 descriptions"]

    descriptions: dict[str, dict] = {}
    current_major: dict[str, str] = {}
    current_sub_major: dict[str, str] = {}
    current_minor: dict[str, str] = {}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        major_code, sub_major_code, minor_code, unit_code = row[:4]
        title = normalize_blank(row[4])
        classified_within = normalize_blank(row[5])
        group_description = normalize_blank(row[6])
        entry_routes = normalize_blank(row[7])
        tasks = split_tilde_bullets(row[8])
        related_job_titles = split_tilde_bullets(row[9])

        if isinstance(major_code, (int, float)):
            current_major = {
                "code": str(int(major_code)),
                "title": title,
                "description": group_description,
            }
        if isinstance(sub_major_code, (int, float)):
            current_sub_major = {
                "code": str(int(sub_major_code)),
                "title": title,
                "description": group_description,
            }
        if isinstance(minor_code, (int, float)):
            current_minor = {
                "code": str(int(minor_code)),
                "title": title,
                "description": group_description,
            }
        if not isinstance(unit_code, (int, float)):
            continue

        code = str(int(unit_code))
        descriptions[code] = {
            "soc_code": code,
            "title": title,
            "group_description": group_description,
            "entry_routes": entry_routes,
            "tasks": tasks,
            "related_job_titles": related_job_titles,
            "classified_within": classified_within,
            "major_group": current_major.copy(),
            "sub_major_group": current_sub_major.copy(),
            "minor_group": current_minor.copy(),
        }

    return descriptions


def render_uk_markdown(occupation: dict, description: dict | None) -> str:
    title = occupation["title"]
    lines = [f"# {title}", ""]
    lines.append(f"**SOC 2020 Unit Group:** {occupation.get('soc_code', 'n/a')}")
    lines.append(f"**Employment Source:** {SOURCE_NAME}")
    lines.append(f"**Description Source:** {DESCRIPTION_SOURCE_NAME}")
    lines.append("")

    if description:
        major = description.get("major_group", {})
        sub_major = description.get("sub_major_group", {})
        minor = description.get("minor_group", {})

        lines.append("## Classification Context")
        lines.append("")
        lines.append(f"- **Major group:** {major.get('code', '')} {major.get('title', '').strip()}".strip())
        if sub_major.get("title"):
            lines.append(
                f"- **Sub-major group:** {sub_major.get('code', '')} {sub_major.get('title', '').strip()}".strip()
            )
        if minor.get("title"):
            lines.append(
                f"- **Minor group:** {minor.get('code', '')} {minor.get('title', '').strip()}".strip()
            )
        lines.append("")

        if description.get("group_description"):
            lines.append("## Occupation Description")
            lines.append("")
            lines.append(description["group_description"])
            lines.append("")

        if description.get("entry_routes"):
            lines.append("## Typical Entry Routes And Qualifications")
            lines.append("")
            lines.append(description["entry_routes"])
            lines.append("")

        if description.get("tasks"):
            lines.append("## Typical Tasks")
            lines.append("")
            for task in description["tasks"]:
                lines.append(f"- {task}")
            lines.append("")

        if description.get("related_job_titles"):
            lines.append("## Related Job Titles")
            lines.append("")
            for related in description["related_job_titles"]:
                lines.append(f"- {related}")
            lines.append("")

        broader_sections = []
        if major.get("description"):
            broader_sections.append(("Major group", major["description"]))
        if sub_major.get("description"):
            broader_sections.append(("Sub-major group", sub_major["description"]))
        if minor.get("description"):
            broader_sections.append(("Minor group", minor["description"]))
        if description.get("classified_within"):
            broader_sections.append(("Classification structure", description["classified_within"]))

        if broader_sections:
            lines.append("## Broader Group Context")
            lines.append("")
            for heading, text in broader_sections:
                lines.append(f"### {heading}")
                lines.append("")
                lines.append(text)
                lines.append("")
    else:
        lines.append("## Occupation Description")
        lines.append("")
        lines.append(
            "No SOC 2020 description was found in the local ONS description workbook for this code."
        )
        lines.append("")

    lines.append("## Labour Market Snapshot")
    lines.append("")
    lines.append(
        f"- **2024 reported employment (lower bound):** {format_number(occupation.get('num_jobs_2024'))}"
    )
    lines.append(
        f"- **2021 to 2024 trend:** {format_percent(occupation.get('trend_pct_2021_2024'))} ({occupation.get('trend_desc', 'n/a')})"
    )
    lines.append(
        f"- **Employment change, 2021 to 2024:** {format_number(occupation.get('employment_change_2021_2024'))}"
    )
    lines.append(
        f"- **Dominant 2024 industry:** {occupation.get('dominant_industry_code', '')} {occupation.get('dominant_industry', '')}".strip()
    )
    if occupation.get("dominant_industry_share") is not None:
        lines.append(
            f"- **Dominant industry share of reported employment:** {occupation['dominant_industry_share'] * 100:.1f}%"
        )
    lines.append(
        f"- **Suppressed 2024 cells in APS workbook:** {occupation.get('suppressed_cells_2024', 0)}"
    )
    lines.append("")

    top_industries = occupation.get("top_industries_2024") or []
    if top_industries:
        lines.append("## Top 2024 Industries")
        lines.append("")
        for item in top_industries:
            lines.append(f"- **{item['code']} {item['label']}**: {format_number(item['jobs'])}")
        lines.append("")

    lines.append("---")
    lines.append(
        "*Note: APS occupation totals here are lower-bound sums of reported industry cells. "
        "ONS suppresses small cells with `*`, so some employment is omitted from the numeric total.*"
    )
    lines.append("")

    return "\n".join(lines)
